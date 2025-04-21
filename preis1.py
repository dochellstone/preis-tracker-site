from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import time
from datetime import datetime
import re

# 🔧 EdgeDriver-Pfad und Pfad zur Link-Datei
EDGEDRIVER_PATH = r"C:\Users\Dr Hellstone\Downloads\edgedriver_win64\msedgedriver.exe"
LINKDATEI_PFAD = r"C:\Users\Dr Hellstone\Documents\artikel_links.txt"
EXCEL_DATEI = "preis_tracking.xlsx"

# Links laden
with open(LINKDATEI_PFAD, "r", encoding="utf-8") as f:
    urls = [line.strip() for line in f if line.strip()]

# Edge Browser im Hintergrund starten
options = webdriver.EdgeOptions()
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--log-level=3")  # Nur Fehler anzeigen
service = Service(EDGEDRIVER_PATH)
driver = webdriver.Edge(service=service, options=options)

# Excel vorbereiten
if os.path.exists(EXCEL_DATEI):
    workbook = load_workbook(EXCEL_DATEI)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preise"
    sheet.append(["Datum", "Produkt", "Preis (€)", "Änderung (€)", "Bester Preis (€)", "Link"])
    bold_font = Font(bold=True)
    for col in range(1, 7):
        sheet.cell(row=1, column=col).font = bold_font

# Aktuelles Datum
datum = datetime.now().strftime("%Y-%m-%d")

# Produkte abrufen
for url in urls:
    driver.get(url)
    time.sleep(3)

    try:
        # Preis-Element je nach Domain
        if "otto.de" in url:
            price_element = driver.find_element(By.CLASS_NAME, "js_pdp_price__retail-price__value_")

        elif "mediamarkt" in url:
            price_element = driver.find_element(By.CSS_SELECTOR, '[itemprop="price"]')

        else:
            price_element = driver.find_element(By.CLASS_NAME, "product-detail-price")

        # Preis bereinigen
        raw_price = price_element.text.strip()
        price_text = re.sub(r"[^\d,\.]", "", raw_price).replace(".", "").replace(",", ".")
        price = float(price_text)

        # Produktname
        product_name = driver.title.strip().replace(" | MediaMarkt", "").replace(" jetzt online kaufen", "")

        # Vorherige Preise prüfen
        previous_prices = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == product_name:
                previous_prices.append(row[2])

        # Preisänderung berechnen
        last_price = previous_prices[-1] if previous_prices else None
        change = round(price - last_price, 2) if last_price is not None else None

        # Bester Preis berechnen
        best_price = min(previous_prices + [price]) if previous_prices else price

        # Neue Zeile schreiben
        row_data = [datum, product_name, price, change if change is not None else "–", best_price, url]
        sheet.append(row_data)

        # Hyperlink einfügen
        link_cell = sheet.cell(row=sheet.max_row, column=6)
        link_cell.hyperlink = url
        link_cell.style = "Hyperlink"

        print(f"✅ {product_name}: {price} € gespeichert.")

    except Exception as e:
        print(f"❌ Fehler bei {url}: {e}")

# Summen einfügen (z. B. 5 Zeilen Abstand nach den Daten)
sum_row = sheet.max_row + 5
sheet[f"B{sum_row}"] = "Gesamtpreis:"
sheet[f"B{sum_row}"].font = Font(bold=True)

# Preise summieren (Spalte C)
preis_spalte = "C"
startzeile = 2
endzeile = sheet.max_row - 5
sheet[f"C{sum_row}"] = f"=SUM({preis_spalte}{startzeile}:{preis_spalte}{endzeile})"
sheet[f"C{sum_row}"].font = Font(bold=True)

# Auto-Spaltenbreite
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

# Speichern
workbook.save(EXCEL_DATEI)
driver.quit()
