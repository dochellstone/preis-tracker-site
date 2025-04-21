from openpyxl import load_workbook
import os

EXCEL_DATEI = "preis_tracking.xlsx"
HTML_DATEI = "preis_tracking.html"

# 🔍 Existenz prüfen
if not os.path.exists(EXCEL_DATEI):
    print(f"❌ Excel-Datei '{EXCEL_DATEI}' nicht gefunden!")
    exit(1)

# Excel-Datei laden
workbook = load_workbook(EXCEL_DATEI)
sheet = workbook.active

# HTML-Basisstruktur
html = """
<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <title>Preisübersicht</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 30px; background-color: #f4f4f4; }
        h1 { color: #333; }
        table { border-collapse: collapse; width: 100%; background: white; }
        th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
        th { background-color: #eee; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        a { color: #0066cc; text-decoration: none; }
    </style>
</head>
<body>
    <h1>🛒 Preisübersicht</h1>
    <table>
        <tr>
"""

# Tabellenkopf aus erster Zeile
for cell in sheet[1]:
    html += f"<th>{cell.value}</th>"
html += "</tr>\n"

# Datenzeilen
for row in sheet.iter_rows(min_row=2, values_only=True):
    if all(cell is None for cell in row):
        continue  # leere Zeilen überspringen

    html += "<tr>"
    current_price = None
    best_price = None

    for i, cell in enumerate(row):
        if i == 2:  # Preis (€)
            try:
                current_price = float(cell)
                html += f"<td>{current_price:.2f}</td>"
            except:
                html += f"<td>{cell if cell is not None else ''}</td>"

        elif i == 3:  # Änderung (€)
            try:
                diff = float(cell)
                color = "red" if diff < 0 else "blue"
                symbol = "🔻" if diff < 0 else "🔺"
                html += f'<td style="color:{color};">{symbol} {abs(diff):.2f}</td>'
            except:
                html += f"<td>{cell if cell is not None else ''}</td>"

        elif i == 4:  # Bester Preis
            try:
                best_price = float(cell)
                if current_price is not None and best_price == current_price:
                    html += f'<td style="background-color:#d4edda; font-weight:bold;">{best_price:.2f}</td>'
                else:
                    html += f"<td>{best_price:.2f}</td>"
            except:
                html += f"<td>{cell if cell is not None else ''}</td>"

        elif i == 5 and isinstance(cell, str):  # Link
            html += f'<td><a href="{cell}" target="_blank">🔗 Link</a></td>'

        else:
            html += f"<td>{cell if cell is not None else ''}</td>"

    html += "</tr>\n"

html += """
    </table>
</body>
</html>
"""

# HTML-Datei speichern
with open(HTML_DATEI, "w", encoding="utf-8") as f:
    f.write(html)

print(f"✅ HTML-Datei erfolgreich erstellt: {HTML_DATEI}")
